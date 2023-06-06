import os.path

import cv2
import numpy as np
import pyzbar.pyzbar as pyzbar
from PIL import Image as PilImg

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from io import BytesIO
import math
import pytesseract
from reportlab.lib.pagesizes import letter

from config import DEFAULT_QR
from forcedisplays import *
import qrcode
from reportlab.pdfgen import canvas
from api.models import Item


def recognize_qr(file):
    inp = np.asarray(bytearray(file), dtype=np.uint8)
    img = cv2.imdecode(inp, cv2.IMREAD_COLOR)

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    cv2.imwrite('gray.jpg', gray)

    edges = cv2.Canny(gray, 50, 150)
    cv2.imwrite('edges.jpg', edges)

    contours, hierarchy = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

    for contour in contours:
        approx = cv2.approxPolyDP(contour, 0.01 * cv2.arcLength(contour, True), True)
        if len(approx) == 4:
            rect = cv2.minAreaRect(contour)
            box = cv2.boxPoints(rect)
            box = np.int0(box)
            cv2.drawContours(img, [box], 0, (0, 0, 255), 2)

    decoded_objects = pyzbar.decode(gray)

    content = None
    for obj in decoded_objects:
        content = obj.data.decode('utf-8')
        print("Data: ", content)
    return content


def get_table_display():
    display = {key: forces['item'][key] for key in table_display if key in forces['item']}

    return display


def get_qr_display(is_computer):
    display = {key: forces['item'][key] for key in qr_display if key in forces['item']}
    if is_computer:
        display.update(forces['computer'])

    return display


def generate_xlsx(items):
    workbook = Workbook()
    worksheet = workbook.active
    display = get_table_display()

    header = ['Номер'] + list(display.values()) + ['Ответственный']
    for col_num, header_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f'{col_letter}1'] = header_title

    for row_num, item in enumerate(items, start=2):
        worksheet[f'A{row_num}'] = row_num - 1
        for col_num, key in enumerate(display.keys(), start=2):
            col_letter = get_column_letter(col_num)
            worksheet[f'{col_letter}{row_num}'] = item['item'].get(key, '')

        worksheet[f'{get_column_letter(len(display) + 2)}{row_num}'] = item['user']

    excel_file = BytesIO()

    # Save the workbook to the BytesIO object
    workbook.save(excel_file)
    excel_file.seek(0)

    return excel_file


def generate_movement_xlsx(movement, items):
    workbook = Workbook()
    worksheet = workbook.active
    display = get_table_display()

    worksheet[f'A1'] = 'Сотруднику'
    worksheet[f'B1'] = movement.username
    worksheet[f'A2'] = 'Дата'
    worksheet[f'B2'] = movement.date.strftime('%d.%m.%Y')
    worksheet[f'A3'] = 'До даты'
    worksheet[f'B3'] = movement.move_date.strftime('%d.%m.%Y')
    worksheet[f'A4'] = 'Местоположение'
    worksheet[f'B4'] = movement.location

    table_header = ['Номер'] + list(display.values()) + ['Ответственный'] + ['Вернули']
    for col_num, header_title in enumerate(table_header, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f'{col_letter}6'] = header_title

    for row_num, item in enumerate(items, start=7):
        worksheet[f'A{row_num}'] = row_num - 6
        for col_num, key in enumerate(display.keys(), start=2):
            col_letter = get_column_letter(col_num)
            worksheet[f'{col_letter}{row_num}'] = item['item'].get(key, '')

        worksheet[f'{get_column_letter(len(display) + 2)}{row_num}'] = item['user']
        worksheet[f'{get_column_letter(len(display) + 3)}{row_num}'] = 'Да' if item['is_returned'] else 'Нет'

    excel_file = BytesIO()

    # Save the workbook to the BytesIO object
    workbook.save(excel_file)
    excel_file.seek(0)

    return excel_file


def generate_qr(items):
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)

    app_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(app_dir, 'qr_template.png')

    template_image = cv2.imread(template_path)
    template_height, template_width, _ = template_image.shape

    qr_width = 120
    qr_height = 120

    pdf_width, pdf_height = letter

    num_columns = int(pdf_width / template_width)
    num_rows = int(pdf_height / template_height)

    x_text, y_text = find_text_position(template_image, template_image)

    num_pages = math.ceil(len(items) / (num_columns * num_rows))

    print(num_pages)

    for page in range(num_pages):
        pdf.showPage()

        for row in range(num_rows):
            for column in range(num_columns):
                index = page * (num_columns * num_rows) + row * num_columns + column
                if index >= len(items):
                    break

                item = items[index]
                db_item = Item.check_or_create_item(item['id'], item['item_type'])

                qr = qrcode.QRCode(version=1, box_size=10, border=1)
                qr.add_data(DEFAULT_QR.format(db_item.item_type, db_item.guid))
                qr.make(fit=True)

                img = qr.make_image(fill_color="black", back_color="white")

                x = column * template_width
                y = pdf_height - (row + 1) * template_height

                x_qr = x + (template_width - qr_width) - 2
                y_qr = y + (template_height - qr_height) // 2

                pdf.drawInlineImage(template_path, x, y, width=template_width, height=template_height)
                pdf.drawInlineImage(img, x_qr, y_qr, width=qr_width, height=qr_height)

    pdf.save()
    buffer.seek(0)
    return buffer


def find_text_position(image, template):
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    gray_template = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
    result = cv2.matchTemplate(gray_image, gray_template, cv2.TM_CCOEFF_NORMED)
    _, _, _, max_loc = cv2.minMaxLoc(result)
    w, h = template.shape[1], template.shape[0]
    x = max_loc[0] + w // 2
    y = max_loc[1] + h // 2
    return x, y
